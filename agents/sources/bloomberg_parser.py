"""
Parser dos dados brutos de notícias do Bloomberg Terminal → CSV do JEMPO.

O João coletou notícias manualmente no Terminal Bloomberg da biblioteca da FGV
(copiar/colar da tela TOP News) e salvou num Excel com uma aba por ticker. Cada
aba traz blocos de notícia completos no formato do Terminal, poluídos por
"chrome" da interface (`<Back> Voltar`, barra de ferramentas, `Painel gráfico »`,
tags `<TICKER> BZ Equity`).

Os blocos ficam normalmente na coluna A, mas parte das abas da 2ª coleta traz
todos (EGIE3) ou alguns (RDOR3, BBSE3) deslocados para a coluna B. Por isso A e
B são parseadas como fluxos independentes e unidas — ver `_parsear_worksheet`.

Abas cujo nome comercial diverge do ticker do `UNIVERSO_HISTORICO` passam por
`TICKER_ALIAS` (hoje só `AXIA3` → `ELET3`, rebrand pós-privatização).

Da coluna D em diante há um índice tabular só-headline com datas corrompidas
(espalhadas de 1930 a 2029) — descartado de propósito: não tem corpo, não está
no formato descrito e envenenaria o corte anti-lookahead. Ele nunca casa com a
âncora `MM/DD/YYYY HH:MM:SS[CODE]`, então não entra pelo caminho de A/B.

Esta é a Etapa 1 (parser + CSV). A integração com o JournalAgent é a Etapa 2,
separada — aqui apenas produzimos o CSV no schema esperado.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# ── Constantes de domínio ─────────────────────────────────────────────────────

_FUSO = "America/Sao_Paulo"          # Terminal mostra horário local do usuário (SP)
_SUFIXO_TICKER = ".SA"               # sufixo yfinance para ações brasileiras
_PESO_BLOOMBERG = 1.0                # fonte primária no JEMPO
_URL_VAZIA = ""                      # Bloomberg não fornece URL pública
_MIN_CHARS_TITULO = 10               # títulos abaixo disso são chrome/lixo

# Formato do timestamp bruto do Terminal: MM/DD/YYYY HH:MM:SS
_FMT_TIMESTAMP = "%m/%d/%Y %H:%M:%S"

# Ordem exata das colunas do CSV consumido pelo JEMPO.
_COLUNAS_CSV = ("data", "ticker", "titulo", "fonte", "url", "peso", "corpo", "resumo_ia")

# Mapeamento código do Terminal → nome da fonte. Códigos fora daqui viram
# "Bloomberg" genérico e são reportados no relatório (regra do gate de custo).
_MAPA_FONTE = {
    "BN": "Bloomberg News",
    "BFW": "Bloomberg First Word",
    "PBN": "Bloomberg Portuguese",
    "BI": "Bloomberg Intelligence",
}
_FONTE_GENERICA = "Bloomberg"

# Nome de aba do Terminal → ticker do UNIVERSO_HISTORICO, quando divergem.
# O universo é a fonte de verdade; remapear aqui mantém a correção localizada
# e reversível, sem tocar no histórico 2019-2023 já treinado.
TICKER_ALIAS = {
    "AXIA3": "ELET3",  # rebrand pós-privatização (2024); universo JEMPO usa o
                       # nome histórico ELET3 para consistência com 2019-2023
}

# ── Regex compilados (fora dos loops, por performance) ────────────────────────

# Linha de data que abre cada notícia: " MM/DD/YYYY HH:MM:SS[CODE]".
# Usada como âncora de segmentação (mais robusta que <Back> Voltar, que a
# colagem do índice tabular às vezes engole).
_RE_DATA_FONTE = re.compile(
    r"(?P<data>\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})\[(?P<codigo>\w+)\]"
)

# Marcadores de início de corpo.
_RE_BLOOMBERG_DASH = re.compile(r"\(Bloomberg\)\s*--")
_RE_BY_AUTOR = re.compile(r"^\s*By\s+\S", re.IGNORECASE)

# Seções especiais.
_RE_RESUMO_IA = re.compile(r"Resumo da Bloomberg AI", re.IGNORECASE)
_RE_BULLET = re.compile(r"^\s*[•·]\s*")

# Linhas a ignorar como chrome puro.
_RE_BACK_VOLTAR = re.compile(r"^\s*<Back>\s*Voltar\s*$", re.IGNORECASE)
_RE_TOOLBAR = re.compile(r"^\s*Anterior\s+Pr[óo]xima", re.IGNORECASE)
_RE_ATUALIZADA = re.compile(r"^\s*Esta not[íi]cia foi atualizada", re.IGNORECASE)
_RE_SUBSCRIBE = re.compile(r"^\s*Subscribe\b", re.IGNORECASE)

# Chrome embutido inline no meio do texto (removido, mantendo o resto da linha).
_RE_PAINEL_GRAFICO = re.compile(r"\s*Painel gr[áa]fico\s*»?")
_RE_BZ_EQUITY = re.compile(r"\s*[A-Z0-9]{4,6}\s+BZ Equity\s*")

# Chrome residual no fim do corpo (rodapé padrão do Terminal). Só afeta `corpo`.
# Marcadores de rodapé: a partir do 1º match, tudo é boilerplate (créditos do
# repórter/editor/tradutor, e-mails, bloco "Ativos", "Notícias recomendadas").
_RE_RODAPE_CORPO = re.compile(
    r"Para entrar em contato com"
    r"|To contact the (?:reporter|reporters|editor|editors|translator|translators)"
    r"|Not[íi]cias recomendadas",
    re.IGNORECASE,
)
# Tag de ativo "TICKER BZ" sem "Equity" (ex: "BBDC4 BZ"); formato inequívoco de
# ticker Bloomberg — improvável em texto real. Removida inline.
_RE_TAG_TICKER_BZ = re.compile(r"\s*\b[A-Z]{4}\d+ BZ\b")
# Run FINAL de referências numéricas do menu TOP News (ex: "101) 102) 103)").
# Exige 2+ tokens para não apagar um parêntese solto de conteúdo real
# (data "(11/26/25)" ou preço-alvo "R$ 42)").
_RE_NUMS_MENU_TRAILING = re.compile(r"(?:\s*\d{2,3}\)){2,}\s*$")


# ── Contratos públicos ────────────────────────────────────────────────────────


@dataclass(frozen=True)
class NoticiaBloomberg:
    """Uma notícia Bloomberg tipada, pronta para o schema do JEMPO."""

    data: pd.Timestamp   # tz-aware America/Sao_Paulo
    ticker: str          # com sufixo .SA
    titulo: str
    fonte: str
    url: str             # "" para Bloomberg
    peso: float          # 1.0
    corpo: str
    resumo_ia: str


@dataclass(frozen=True)
class RelatorioParsing:
    """Sumário auditável do parsing de todo o Excel."""

    n_noticias_extraidas: int
    n_por_ticker: dict[str, int]
    abas_vazias: list[str]
    codigos_fonte_desconhecidos: list[str]
    n_duplicatas_removidas: int
    n_puladas_titulo_vazio: int
    n_puladas_data_invalida: int
    avisos: list[str]


@dataclass
class _RelatorioAba:
    """Contadores parciais de uma única aba (agregados no relatório final)."""

    codigos_fonte_desconhecidos: list[str] = field(default_factory=list)
    n_duplicatas_removidas: int = 0
    n_puladas_titulo_vazio: int = 0
    n_puladas_data_invalida: int = 0
    avisos: list[str] = field(default_factory=list)


# ── Helpers puros ──────────────────────────────────────────────────────────────


def ticker_da_aba(aba: str) -> str:
    """Nome da aba → ticker com sufixo .SA (`petr4` → `PETR4.SA`).

    Aplica `TICKER_ALIAS` antes do sufixo, para abas que usam um nome comercial
    diferente do adotado no universo (`AXIA3` → `ELET3.SA`).
    """
    nome = aba.strip().upper()
    return f"{TICKER_ALIAS.get(nome, nome)}{_SUFIXO_TICKER}"


def mapear_codigo_fonte(codigo: str) -> tuple[str, bool]:
    """Código do Terminal → (nome da fonte, código_conhecido?).

    Códigos fora do mapa viram "Bloomberg" genérico e retornam False no
    segundo elemento, para o chamador registrar como desconhecido.
    """
    nome = _MAPA_FONTE.get(codigo.upper())
    if nome is None:
        return _FONTE_GENERICA, False
    return nome, True


def parsear_timestamp_bloomberg(raw: str) -> Optional[pd.Timestamp]:
    """`MM/DD/YYYY HH:MM:SS` (naive, horário SP) → Timestamp tz-aware SP.

    Retorna None se a data não parsear (mês/dia/hora impossíveis, formato
    inesperado), para o chamador pular a notícia e logar.
    """
    try:
        naive = pd.to_datetime(raw.strip(), format=_FMT_TIMESTAMP)
    except (ValueError, TypeError):
        return None
    return naive.tz_localize(_FUSO)


# ── Núcleo do parsing (texto → notícias) ───────────────────────────────────────


def _eh_chrome(linha: str) -> bool:
    """True se a linha é puro chrome de interface (a descartar por inteiro)."""
    return bool(
        _RE_BACK_VOLTAR.match(linha)
        or _RE_TOOLBAR.match(linha)
    )


def _limpar_inline(linha: str) -> str:
    """Remove chrome embutido no meio da linha, preservando o texto real."""
    linha = _RE_PAINEL_GRAFICO.sub(" ", linha)
    linha = _RE_BZ_EQUITY.sub(" ", linha)
    return linha


def _normalizar_texto(linhas: list[str]) -> str:
    """Junta linhas de corpo/resumo limpas num blob de espaços únicos."""
    limpas = (_limpar_inline(l) for l in linhas if not _eh_chrome(l))
    return re.sub(r"\s+", " ", " ".join(limpas)).strip()


def _limpar_chrome_residual(corpo: str) -> str:
    """Remove chrome residual do rodapé do corpo (só o campo `corpo`).

    Trunca no primeiro marcador de rodapé (créditos de contato ou "Notícias
    recomendadas"), remove tags "TICKER BZ" e o run final de referências do menu
    ("101) 102) 103)"). Preserva conteúdo real: datas em parênteses, preços-alvo
    e cross-refs "NOTE:"/"RELATED:" ficam intactos.
    """
    if not corpo:
        return corpo
    m = _RE_RODAPE_CORPO.search(corpo)
    if m:
        corpo = corpo[: m.start()]
    corpo = _RE_TAG_TICKER_BZ.sub("", corpo)
    corpo = _RE_NUMS_MENU_TRAILING.sub("", corpo)
    return re.sub(r"\s+", " ", corpo).strip()


def _segmentar(texto: str) -> list[list[str]]:
    """Quebra o texto da aba em blocos, um por notícia.

    Âncora: a linha de data `MM/DD/YYYY HH:MM:SS[CODE]`. Cada bloco vai de uma
    linha-de-data (inclusive) até imediatamente antes da próxima. O que vier
    antes da primeira linha-de-data é preâmbulo descartado.
    """
    blocos: list[list[str]] = []
    atual: Optional[list[str]] = None
    for linha in texto.splitlines():
        if _RE_DATA_FONTE.search(linha):
            atual = [linha]
            blocos.append(atual)
        elif atual is not None:
            atual.append(linha)
    return blocos


def _extrair_titulo(linhas_pos_data: list[str]) -> tuple[Optional[str], int]:
    """Primeira linha significativa após a data → (título, índice da linha).

    Ignora chrome, `Esta notícia foi atualizada`, `Subscribe`, rótulo/bullets do
    resumo e marcadores de corpo. Título precisa ter >= _MIN_CHARS_TITULO chars.
    O índice (na lista `linhas_pos_data`) permite ao corpo, na falta de
    marcadores, começar logo após o título. Retorna (None, -1) se não houver.
    """
    for i, linha in enumerate(linhas_pos_data):
        stripped = linha.strip()
        if not stripped:
            continue
        if (
            _eh_chrome(linha)
            or _RE_ATUALIZADA.match(linha)
            or _RE_SUBSCRIBE.match(linha)
            or _RE_RESUMO_IA.search(linha)
            or _RE_BULLET.match(linha)
            or _RE_BY_AUTOR.match(linha)
            or _RE_BLOOMBERG_DASH.search(linha)
        ):
            continue
        candidato = _limpar_inline(stripped).strip()
        # Uma linha que era só a tag "<TICKER> BZ Equity" fica vazia após limpar.
        if len(candidato) >= _MIN_CHARS_TITULO:
            return candidato, i
        # Linha significativa porém curta → título inválido (não seguir procurando).
        return (candidato or None), i
    return None, -1


def _extrair_resumo(linhas_pos_data: list[str]) -> str:
    """Texto do bloco `Resumo da Bloomberg AI` (bullets), ou "" se ausente.

    Captura tudo entre o rótulo do resumo e o início do corpo (`By ` ou
    `(Bloomberg) --`), retirando marcadores de bullet.
    """
    coletando = False
    partes: list[str] = []
    for linha in linhas_pos_data:
        if _RE_RESUMO_IA.search(linha):
            coletando = True
            continue
        if not coletando:
            continue
        if _RE_BY_AUTOR.match(linha) or _RE_BLOOMBERG_DASH.search(linha):
            break
        partes.append(_RE_BULLET.sub("", linha))
    return _normalizar_texto(partes)


def _extrair_corpo(linhas_pos_data: list[str], idx_titulo: int) -> str:
    """Corpo da notícia a partir de `(Bloomberg) --` ou, na falta, de `By `.

    Sem esses marcadores (comum em itens traduzidos, código VAL), o corpo é o
    texto que segue o título — evitando descartar a matéria inteira. Retorna ""
    apenas se não houver nenhuma linha de conteúdo após o título.
    """
    inicio = None
    for i, linha in enumerate(linhas_pos_data):
        if _RE_BLOOMBERG_DASH.search(linha):
            inicio = i
            break
    if inicio is None:
        for i, linha in enumerate(linhas_pos_data):
            if _RE_BY_AUTOR.match(linha):
                inicio = i + 1  # o corpo começa após a linha do autor
                break
    if inicio is None:
        # Fallback: sem marcadores, o corpo é o que vem depois do título.
        inicio = idx_titulo + 1 if idx_titulo >= 0 else 0
    return _limpar_chrome_residual(_normalizar_texto(linhas_pos_data[inicio:]))


def parsear_texto_aba(texto: str, aba: str) -> tuple[list[NoticiaBloomberg], _RelatorioAba]:
    """Parseia o texto bruto da coluna A de uma aba em NoticiaBloomberg.

    Deduplica por (data, titulo) dentro da aba. Retorna as notícias e os
    contadores parciais (código desconhecido, puladas, duplicatas) da aba.
    """
    ticker = ticker_da_aba(aba)
    rel = _RelatorioAba()
    noticias: list[NoticiaBloomberg] = []
    vistas: set[tuple[str, str]] = set()

    for bloco in _segmentar(texto):
        m = _RE_DATA_FONTE.search(bloco[0])
        if m is None:
            continue  # não deveria acontecer: _segmentar só abre bloco com match

        data = parsear_timestamp_bloomberg(m.group("data"))
        if data is None:
            rel.n_puladas_data_invalida += 1
            logger.warning("%s: data inválida %r; pulando notícia", aba, m.group("data"))
            continue

        fonte, conhecido = mapear_codigo_fonte(m.group("codigo"))
        if not conhecido and m.group("codigo") not in rel.codigos_fonte_desconhecidos:
            rel.codigos_fonte_desconhecidos.append(m.group("codigo"))
            logger.warning("%s: código de fonte desconhecido %r → %s",
                           aba, m.group("codigo"), _FONTE_GENERICA)

        linhas_pos_data = bloco[1:]
        titulo, idx_titulo = _extrair_titulo(linhas_pos_data)
        if titulo is None or len(titulo) < _MIN_CHARS_TITULO:
            rel.n_puladas_titulo_vazio += 1
            logger.warning("%s: título vazio/curto (%r); pulando notícia", aba, titulo)
            continue

        corpo = _extrair_corpo(linhas_pos_data, idx_titulo)
        resumo = _extrair_resumo(linhas_pos_data)
        if not corpo and not resumo:
            logger.warning("%s: notícia sem corpo nem resumo (%r); mantendo", aba, titulo)

        chave = (data.isoformat(), titulo)
        if chave in vistas:
            rel.n_duplicatas_removidas += 1
            logger.warning("%s: duplicata (data,titulo) removida: %r", aba, titulo)
            continue
        vistas.add(chave)

        noticias.append(NoticiaBloomberg(
            data=data,
            ticker=ticker,
            titulo=titulo,
            fonte=fonte,
            url=_URL_VAZIA,
            peso=_PESO_BLOOMBERG,
            corpo=corpo,
            resumo_ia=resumo,
        ))

    return noticias, rel


# ── Escrita do CSV ─────────────────────────────────────────────────────────────


def _chave_ordenacao(n: NoticiaBloomberg) -> tuple[str, str, str]:
    """Ordenação estável (determinismo byte-a-byte entre runs)."""
    return (n.ticker, n.data.isoformat(), n.titulo)


def escrever_csv(noticias: list[NoticiaBloomberg], caminho: Path) -> None:
    """Grava as notícias no CSV UTF-8 com o schema exato do JEMPO.

    Ordenação estável por (ticker, data, titulo) garante saída idêntica em
    execuções repetidas com a mesma entrada.
    """
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    ordenadas = sorted(noticias, key=_chave_ordenacao)
    df = pd.DataFrame(
        [
            {
                "data": n.data.isoformat(),
                "ticker": n.ticker,
                "titulo": n.titulo,
                "fonte": n.fonte,
                "url": n.url,
                "peso": n.peso,
                "corpo": n.corpo,
                "resumo_ia": n.resumo_ia,
            }
            for n in ordenadas
        ],
        columns=list(_COLUNAS_CSV),
    )
    df.to_csv(caminho, index=False, encoding="utf-8", lineterminator="\n")


# ── Orquestração: Excel → notícias + CSV ───────────────────────────────────────


def _texto_coluna(worksheet, indice: int) -> str:
    """Concatena as células não-vazias de uma coluna (0=A, 1=B) com `\\n`."""
    col = indice + 1
    linhas = [
        str(cell.value)
        for (cell,) in worksheet.iter_rows(min_col=col, max_col=col)
        if cell.value not in (None, "")
    ]
    return "\n".join(linhas)


def _detectar_coluna_conteudo(worksheet) -> int:
    """Índice da coluna com mais date-lines: 0 (A), 1 (B), ou -1 se nenhuma.

    Parte das abas exportadas do Terminal traz os blocos deslocados uma coluna
    à direita. A âncora `MM/DD/YYYY HH:MM:SS[CODE]` distingue conteúdo real do
    índice tabular só-headline (que fica da coluna D em diante e nunca casa com
    a âncora), então contá-la é seguro para escolher a coluna.
    """
    contagens = [
        len(_RE_DATA_FONTE.findall(_texto_coluna(worksheet, i)))
        for i in (0, 1)
    ]
    if max(contagens) == 0:
        return -1
    return contagens.index(max(contagens))


def _parsear_worksheet(worksheet, aba: str) -> tuple[list[NoticiaBloomberg], _RelatorioAba]:
    """Parseia as colunas A e B da aba, unindo os resultados.

    As duas colunas são parseadas como fluxos independentes porque abas reais
    trazem blocos divididos entre A e B (o Terminal desloca parte da colagem).
    Escolher só a coluna dominante descartaria notícias legítimas da outra. Uma
    coluna sem âncoras não abre bloco algum e contribui com zero notícias, de
    modo que o índice tabular residual continua fora do resultado.
    """
    noticias: list[NoticiaBloomberg] = []
    rel_total = _RelatorioAba()
    vistas: set[tuple[str, str]] = set()

    for indice in (0, 1):
        parciais, rel = parsear_texto_aba(_texto_coluna(worksheet, indice), aba)
        for n in parciais:
            chave = (n.data.isoformat(), n.titulo)
            if chave in vistas:
                rel_total.n_duplicatas_removidas += 1
                continue
            vistas.add(chave)
            noticias.append(n)
        for c in rel.codigos_fonte_desconhecidos:
            if c not in rel_total.codigos_fonte_desconhecidos:
                rel_total.codigos_fonte_desconhecidos.append(c)
        rel_total.n_duplicatas_removidas += rel.n_duplicatas_removidas
        rel_total.n_puladas_titulo_vazio += rel.n_puladas_titulo_vazio
        rel_total.n_puladas_data_invalida += rel.n_puladas_data_invalida
        rel_total.avisos.extend(rel.avisos)

    return noticias, rel_total


def parsear_excel_bloomberg(
    caminho_excel: Path,
    caminho_csv_saida: Optional[Path] = None,
) -> tuple[list[NoticiaBloomberg], RelatorioParsing]:
    """Parseia o Excel bruto do Bloomberg em NoticiaBloomberg tipada.

    Se `caminho_csv_saida` for fornecido, grava o CSV no schema do JEMPO.
    Retorna as notícias e um relatório auditável agregado por todas as abas.
    """
    caminho_excel = Path(caminho_excel)
    wb = openpyxl.load_workbook(caminho_excel, data_only=True, read_only=True)

    todas: list[NoticiaBloomberg] = []
    n_por_ticker: dict[str, int] = {}
    abas_vazias: list[str] = []
    codigos_desconhecidos: list[str] = []
    n_duplicatas = 0
    n_puladas_titulo = 0
    n_puladas_data = 0
    avisos: list[str] = []

    for aba in wb.sheetnames:
        ws = wb[aba]
        if _detectar_coluna_conteudo(ws) == -1:
            abas_vazias.append(aba)
            logger.warning("Aba %r sem date-lines em A nem B; pulando", aba)
            continue

        noticias, rel = _parsear_worksheet(ws, aba)

        todas.extend(noticias)
        n_por_ticker[ticker_da_aba(aba)] = len(noticias)
        for c in rel.codigos_fonte_desconhecidos:
            if c not in codigos_desconhecidos:
                codigos_desconhecidos.append(c)
        n_duplicatas += rel.n_duplicatas_removidas
        n_puladas_titulo += rel.n_puladas_titulo_vazio
        n_puladas_data += rel.n_puladas_data_invalida
        avisos.extend(rel.avisos)

    wb.close()

    relatorio = RelatorioParsing(
        n_noticias_extraidas=len(todas),
        n_por_ticker=n_por_ticker,
        abas_vazias=abas_vazias,
        codigos_fonte_desconhecidos=codigos_desconhecidos,
        n_duplicatas_removidas=n_duplicatas,
        n_puladas_titulo_vazio=n_puladas_titulo,
        n_puladas_data_invalida=n_puladas_data,
        avisos=avisos,
    )
    logger.info(
        "Bloomberg parser: %d notícias | por ticker=%s | abas vazias=%s | "
        "códigos desconhecidos=%s | duplicatas=%d | puladas(título=%d, data=%d)",
        relatorio.n_noticias_extraidas, n_por_ticker, abas_vazias,
        codigos_desconhecidos, n_duplicatas, n_puladas_titulo, n_puladas_data,
    )

    if caminho_csv_saida is not None:
        escrever_csv(todas, Path(caminho_csv_saida))
        logger.info("CSV Bloomberg gravado em %s", caminho_csv_saida)

    return todas, relatorio
