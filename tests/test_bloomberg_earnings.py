"""
Testes do BloombergEarningsSource — consenso de earnings do Terminal Bloomberg.

As fixtures constroem Excels sintéticos em `tmp_path` no MESMO formato do
arquivo real coletado na FGV (header na linha 1, linha-lixo "Média dos valores
absolutos", datas MM/DD/YYYY como texto, percentuais como texto com "%" e
sentinela "N.M."). Nenhum teste depende de `data/bloomberg/raw/`.

O anti-lookahead é o invariante crítico: nenhum método pode devolver um
earnings com `data_divulgacao > data_limite`.
"""
from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from agents.sources.bloomberg_earnings import (
    BloombergEarningsSource,
    EarningsBloomberg,
)

# ── Fixtures ─────────────────────────────────────────────────────────────────

# Colunas exatas do Excel real (13), na ordem em que o Terminal exporta.
COLUNAS_REAIS = [
    "Dt divulg", "Per", "Per ref", "C", "Divulgado", "Comp", "Estimado",
    "%Surp", "Guidance", "Surp est%", "Var prç%", "Últ 12M", "P/L",
]

# Linha-lixo que o Terminal insere logo abaixo do header em toda aba.
LINHA_MEDIA = ["Média dos valores absolutos"] + [None] * 12


def linha(
    dt_divulg: str,
    per: str,
    per_ref: str,
    divulgado: float | None,
    comp: float | None,
    estimado: float | None,
    surp: str | None,
    var_preco: str | None = "1.00%",
    pl: float | None = 5.0,
) -> list:
    """Uma linha no layout do Excel real (13 colunas, C/Guidance/Surp est% vazias)."""
    return [
        dt_divulg, per, per_ref, None, divulgado, comp, estimado,
        surp, None, None, var_preco, None, pl,
    ]


def escrever_excel(caminho: Path, abas: dict[str, list[list]]) -> Path:
    """Cria um .xlsx com uma aba por ticker, no formato do Terminal."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for nome_aba, linhas in abas.items():
        ws = wb.create_sheet(title=nome_aba)
        ws.append(COLUNAS_REAIS)
        ws.append(LINHA_MEDIA)
        for l in linhas:
            ws.append(l)
    wb.save(caminho)
    return caminho


# PETR4 real (recortado): dois trimestres divulgados + uma estimativa futura.
LINHAS_PETR4 = [
    linha("05/11/2027", "Q1 27", "03/27", None, None, 2.35, None, None, 3.54),
    linha("11/06/2025", "Q3 25", "09/25", 2.533, 2.540, 1.744, "45.64%", "3.77%", 3.95),
    linha("08/07/2025", "Q2 25", "06/25", 2.154, 2.070, 1.871, "10.64%", "-6.15%", 3.80),
    linha("05/12/2025", "Q1 25", "03/25", 2.790, 2.730, 2.740, "-0.36%", "1.52%", None),
]


@pytest.fixture
def excel_padrao(tmp_path: Path) -> Path:
    return escrever_excel(tmp_path / "earnings.xlsx", {"PETR4": LINHAS_PETR4})


def sp(data: str) -> pd.Timestamp:
    return pd.Timestamp(data, tz="America/Sao_Paulo")


# ── Grupo A — Parsing ────────────────────────────────────────────────────────


def test_parseia_aba_com_dados_completos(excel_padrao: Path):
    src = BloombergEarningsSource(excel_padrao)

    e = src.buscar_ultimo_earnings("PETR4.SA", sp("2025-08-20"))

    assert e == EarningsBloomberg(
        ticker="PETR4.SA",
        data_divulgacao=sp("2025-08-07"),
        periodo="Q2 25",
        periodo_ref=sp("2025-06-30"),
        lpa_estimado=1.871,
        lpa_realizado=2.154,
        lpa_comparavel=2.070,
        surpresa_pct=10.64,
        var_preco_pct=-6.15,
        pl=3.80,
    )


def test_ticker_alias_axia3_vira_elet3(tmp_path: Path):
    # Rebrand pós-privatização: a aba vem AXIA3, o universo indexa por ELET3.
    caminho = escrever_excel(tmp_path / "e.xlsx", {"AXIA3": LINHAS_PETR4})
    src = BloombergEarningsSource(caminho)

    assert src.buscar_ultimo_earnings("ELET3.SA", sp("2025-08-20")) is not None
    assert src.buscar_ultimo_earnings("AXIA3.SA", sp("2025-08-20")) is None


def test_nome_de_aba_com_espaco_sobrando_e_normalizado(tmp_path: Path):
    # O arquivo real traz a aba "LREN3 " (com espaço à direita).
    caminho = escrever_excel(tmp_path / "e.xlsx", {"LREN3 ": LINHAS_PETR4})
    src = BloombergEarningsSource(caminho)

    assert src.buscar_ultimo_earnings("LREN3.SA", sp("2025-08-20")) is not None


def test_aba_vazia_nao_crasha(tmp_path: Path, caplog):
    caminho = escrever_excel(tmp_path / "e.xlsx", {"PETR4": LINHAS_PETR4, "VAZIA3": []})
    src = BloombergEarningsSource(caminho)

    with caplog.at_level(logging.WARNING):
        e = src.buscar_ultimo_earnings("PETR4.SA", sp("2025-08-20"))

    assert e is not None
    assert src.buscar_ultimo_earnings("VAZIA3.SA", sp("2025-08-20")) is None
    assert "VAZIA3" in caplog.text


def test_linha_media_dos_valores_absolutos_e_descartada(excel_padrao: Path):
    # A linha-lixo do Terminal não tem data e jamais pode virar um earnings.
    src = BloombergEarningsSource(excel_padrao)

    df = src._carregar_excel()

    assert len(df) == len(LINHAS_PETR4)
    assert df["data_divulgacao"].notna().all()


def test_linhas_futuras_sem_realizado_preservadas(excel_padrao: Path):
    # Estimativas de 2027 existem no Excel; ficam no DataFrame com realizado
    # nulo e só são barradas pelo corte de data_limite dos métodos de busca.
    src = BloombergEarningsSource(excel_padrao)

    df = src._carregar_excel()
    futura = df[df["periodo"] == "Q1 27"].iloc[0]

    assert futura["lpa_estimado"] == 2.35
    assert pd.isna(futura["lpa_realizado"])
    assert pd.isna(futura["surpresa_pct"])


def test_surpresa_nao_mensuravel_vira_none(tmp_path: Path):
    # "N.M." = sinal do LPA virou; surpresa percentual não tem significado.
    caminho = escrever_excel(tmp_path / "e.xlsx", {"PETR4": [
        linha("02/26/2025", "Q4 24", "12/24", 0.015, -1.32, 1.208, "N.M.", "-3.53%", None),
    ]})
    src = BloombergEarningsSource(caminho)

    e = src.buscar_ultimo_earnings("PETR4.SA", sp("2025-03-10"))

    assert e is not None
    assert e.surpresa_pct is None
    assert e.lpa_realizado == 0.015


def test_coluna_pl_ausente_nao_impede_parsing(tmp_path: Path, caplog):
    # A aba AMER3 do arquivo real vem sem a coluna "P/L" (12 colunas).
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="AMER3")
    ws.append(COLUNAS_REAIS[:-1])
    ws.append(LINHA_MEDIA[:-1])
    ws.append(linha("05/12/2025", "Q1 25", "03/25", 1.0, 1.0, 0.8, "25.00%", "2.00%")[:-1])
    caminho = tmp_path / "e.xlsx"
    wb.save(caminho)

    with caplog.at_level(logging.WARNING):
        e = BloombergEarningsSource(caminho).buscar_ultimo_earnings("AMER3.SA", sp("2025-06-01"))

    assert e is not None
    assert e.pl is None
    assert e.surpresa_pct == 25.00


def test_coluna_essencial_ausente_levanta_erro(tmp_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="PETR4")
    ws.append(["Dt divulg", "Per", "EPS Est"])  # sem Estimado/Divulgado/%Surp
    ws.append(["05/12/2025", "Q1 25", 2.74])
    caminho = tmp_path / "e.xlsx"
    wb.save(caminho)

    with pytest.raises(ValueError, match="Estimado"):
        BloombergEarningsSource(caminho).buscar_ultimo_earnings("PETR4.SA", sp("2025-06-01"))


# ── Grupo B — Busca com anti-lookahead ───────────────────────────────────────


def test_buscar_earnings_proximos_dentro_janela(excel_padrao: Path):
    # Resultado em 07/08/2025 (qui); 12/08 (ter) está a 3 dias úteis depois.
    src = BloombergEarningsSource(excel_padrao)

    e = src.buscar_earnings_proximos("PETR4.SA", sp("2025-08-12"), janela_dias=5)

    assert e is not None
    assert e.periodo == "Q2 25"


def test_buscar_earnings_proximos_fora_janela_retorna_none(excel_padrao: Path):
    # 01/09/2025 está a 17 dias úteis do resultado de 07/08 — fora da janela.
    src = BloombergEarningsSource(excel_padrao)

    assert src.buscar_earnings_proximos("PETR4.SA", sp("2025-09-01"), janela_dias=5) is None


def test_buscar_earnings_proximos_respeita_data_limite(excel_padrao: Path):
    # Véspera da divulgação: o resultado do dia seguinte não pode vazar.
    src = BloombergEarningsSource(excel_padrao)

    e = src.buscar_earnings_proximos("PETR4.SA", sp("2025-08-06"), janela_dias=5)

    assert e is None


def test_buscar_earnings_proximos_inclui_o_proprio_dia_da_divulgacao(excel_padrao: Path):
    src = BloombergEarningsSource(excel_padrao)

    e = src.buscar_earnings_proximos("PETR4.SA", sp("2025-08-07 18:00"), janela_dias=5)

    assert e is not None
    assert e.periodo == "Q2 25"


def test_buscar_ultimo_earnings_retorna_mais_recente_antes_de_data_limite(excel_padrao: Path):
    src = BloombergEarningsSource(excel_padrao)

    assert src.buscar_ultimo_earnings("PETR4.SA", sp("2025-12-31")).periodo == "Q3 25"
    assert src.buscar_ultimo_earnings("PETR4.SA", sp("2025-08-20")).periodo == "Q2 25"
    assert src.buscar_ultimo_earnings("PETR4.SA", sp("2025-05-12")).periodo == "Q1 25"


def test_buscar_ultimo_earnings_nunca_devolve_estimativa_futura(excel_padrao: Path):
    # Sem nenhum resultado divulgado antes do limite → None, nunca a linha 2027.
    src = BloombergEarningsSource(excel_padrao)

    assert src.buscar_ultimo_earnings("PETR4.SA", sp("2025-01-10")) is None


def test_linha_de_calendario_sem_realizado_nao_conta_como_divulgacao(excel_padrao: Path):
    # Passado o limite para depois de 2027, a linha Q1 27 satisfaz o corte de
    # data mas nunca foi divulgada: só tem consenso. O último resultado de
    # verdade continua sendo o Q3 25.
    src = BloombergEarningsSource(excel_padrao)

    e = src.buscar_ultimo_earnings("PETR4.SA", sp("2028-01-01"))

    assert e.periodo == "Q3 25"
    assert e.lpa_realizado is not None


def test_ticker_sem_aba_retorna_none(excel_padrao: Path):
    src = BloombergEarningsSource(excel_padrao)

    assert src.buscar_ultimo_earnings("ASAI3.SA", sp("2025-08-20")) is None
    assert src.buscar_earnings_proximos("ASAI3.SA", sp("2025-08-20")) is None


def test_data_limite_naive_levanta_erro(excel_padrao: Path):
    src = BloombergEarningsSource(excel_padrao)

    with pytest.raises(ValueError, match="timezone-aware"):
        src.buscar_ultimo_earnings("PETR4.SA", pd.Timestamp("2025-08-20"))


# ── Grupo C — Robustez ───────────────────────────────────────────────────────


def test_excel_inexistente_retorna_none(tmp_path: Path, caplog):
    src = BloombergEarningsSource(tmp_path / "nao_existe.xlsx")

    with caplog.at_level(logging.WARNING):
        assert src.buscar_ultimo_earnings("PETR4.SA", sp("2025-08-20")) is None
        assert src.buscar_earnings_proximos("PETR4.SA", sp("2025-08-20")) is None

    assert "nao_existe.xlsx" in caplog.text


def test_determinismo_dois_runs_identicos(excel_padrao: Path):
    a = BloombergEarningsSource(excel_padrao).buscar_ultimo_earnings("PETR4.SA", sp("2025-08-20"))
    b = BloombergEarningsSource(excel_padrao).buscar_ultimo_earnings("PETR4.SA", sp("2025-08-20"))

    assert a == b

    # Instâncias independentes: ordenação e tipos têm que bater linha a linha.
    df_a = BloombergEarningsSource(excel_padrao)._carregar_excel()
    df_b = BloombergEarningsSource(excel_padrao)._carregar_excel()
    pd.testing.assert_frame_equal(df_a, df_b)
